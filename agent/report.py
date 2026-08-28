# -*- coding: utf-8 -*-
"""Excel-отчёт о сверке спецификации с чертежами."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='D9E1F2')
FILLS = {
    'ок': PatternFill('solid', fgColor='C6EFCE'),
    'ок (частично)': PatternFill('solid', fgColor='FFEB9C'),
    'расхождение': PatternFill('solid', fgColor='FFC7CE'),
    'нет на чертежах': PatternFill('solid', fgColor='F4B084'),
    'есть на чертежах, метраж не подписан': PatternFill('solid', fgColor='DDEBF7'),
    'узел: кол-во на 1 узел': PatternFill('solid', fgColor='E4DFEC'),
}


def _base_status(s):
    for suf in (' (по измерению)', ' (по кабельному журналу)',
                ' (по ведомости освещения)', ' (по схемам аппаратов)',
                ' (по подписям на чертежах)', ' (по обозначениям на чертежах)'):
        s = s.replace(suf, '')
    return s


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    return ws


def _pp(pages):
    return ', '.join(str(p) for p in sorted(set(pages)))


def write_report(path, project_name, rows, unrows, pages, notes=''):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Сводка
    ws = wb.create_sheet('Сводка')
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 14
    def n_st(name):
        return sum(1 for r in rows if _base_status(r['status']) == name)
    lines = [
        ('Проект', project_name),
        ('Позиций в спецификации (действующих)', len(rows) and sum(1 for _ in rows) + len(unrows)),
        ('Сверено по маркам', len(rows)),
        ('— совпадает (ок)', n_st('ок')),
        ('— совпадает по одному из источников', n_st('ок (частично)')),
        ('— расхождение', n_st('расхождение')),
        ('— не найдено на чертежах', n_st('нет на чертежах')),
        ('— есть на чертежах, метраж не подписан', n_st('есть на чертежах, метраж не подписан')),
        ('— узел: кол-во указано на 1 узел', n_st('узел: кол-во на 1 узел')),
        ('— из них сверено измерением по геометрии',
         sum(1 for r in rows if r['status'].endswith('(по измерению)'))),
        ('— из них сверено по кабельному журналу',
         sum(1 for r in rows if r['status'].endswith('(по кабельному журналу)'))),
        ('— из них сверено по ведомости освещения',
         sum(1 for r in rows if r['status'].endswith('(по ведомости освещения)'))),
        ('— из них сверено по схемам аппаратов',
         sum(1 for r in rows if r['status'].endswith('(по схемам аппаратов)'))),
        ('— из них сверено по подписям на чертежах',
         sum(1 for r in rows if r['status'].endswith('(по подписям на чертежах)'))),
        ('— из них сверено по позиционным обозначениям',
         sum(1 for r in rows if r['status'].endswith('(по обозначениям на чертежах)'))),
        ('Непроверяемо по чертежам (материалы, м.п. и пр.)', len(unrows)),
    ]
    for i, (k, v) in enumerate(lines, 1):
        ws.cell(i, 1, k).font = Font(name=FONT, bold=(i == 1), size=10)
        ws.cell(i, 2, v).font = Font(name=FONT, size=10)
    if notes:
        c = ws.cell(len(lines) + 2, 1, 'Примечания: ' + notes)
        c.font = Font(name=FONT, italic=True, size=9)
        c.alignment = Alignment(wrap_text=True)

    # --- Сверка
    ws = _sheet(wb, 'Сверка',
                ['Марка (канон.)', 'Наименование по спецификации', 'Секции', 'Ед.',
                 'Кол-во Спец.', 'Кол-во План', 'План: число меток / без множителя',
                 'Кол-во Схема', 'Схема: число меток / без множителя',
                 'Точный источник (журнал/ведомость)', 'Статус',
                 'Листы спец. (стр. PDF)', 'Листы План (стр. PDF)', 'Листы Схема (стр. PDF)'],
                [26, 55, 12, 8, 11, 11, 13, 11, 13, 13, 24, 16, 16, 16])
    order = {'расхождение': 0, 'нет на чертежах': 1,
             'есть на чертежах, метраж не подписан': 2,
             'узел: кол-во на 1 узел': 3, 'ок (частично)': 4, 'ок': 5}
    for r in sorted(rows, key=lambda r: (
            order.get(_base_status(r['status']), 5), r['mark'])):
        row = [r['mark'], r['names'], ', '.join(r['sections']), r['unit'],
               r['spec_qty'], r['plan_qty'], r['plan_raw'], r['schema_qty'], r['schema_raw'],
               r.get('journal_qty', ''),
               r['status'], _pp(r['spec_pages']), _pp(r['plan_pages']), _pp(r['schema_pages'])]
        ws.append(row)
        i = ws.max_row
        for j in range(1, len(row) + 1):
            c = ws.cell(i, j)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(wrap_text=(j == 2), vertical='top')
        ws.cell(i, 11).fill = FILLS.get(_base_status(r['status']), HDR_FILL)

    # --- Непроверяемые
    ws = _sheet(wb, 'Непроверяемые',
                ['Стр. PDF', 'Секция', 'Раздел', 'Наименование', 'Марка', 'Ед.', 'Кол-во', 'Причина'],
                [9, 11, 30, 55, 26, 8, 10, 30])
    for it in unrows:
        u = (it['unit'] or '').strip().lower()
        if it.get('uncheck_reason'):
            reason = it['uncheck_reason']
        elif u in ('шт.', 'шт', 'компл.', 'компл', 'к-т'):
            reason = 'марка отсутствует или непригодна для поиска'
        elif u in ('м', 'м.', 'м.п.', 'км'):
            reason = 'нет марки/типоразмера для поиска на чертежах'
        else:
            reason = 'единица измерения не сверяема по чертежам (м², м³, кг и пр.)'
        ws.append([it['page'], it['section'], it['category'], it['name'],
                   it['mark'], it['unit'], it['qty'], reason])
        for j in range(1, 9):
            c = ws.cell(ws.max_row, j)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(wrap_text=(j == 4), vertical='top')

    # --- Листы
    ws = _sheet(wb, 'Классификация листов',
                ['Стр. PDF', 'Тип', 'Название листа'], [9, 12, 80])
    for p in pages:
        ws.append([p['page'], p['kind'], p['title']])
        for j in range(1, 4):
            ws.cell(ws.max_row, j).font = Font(name=FONT, size=9)

    wb.save(path)
    return path
