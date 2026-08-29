# -*- coding: utf-8 -*-
"""Выгрузка в Excel — в тех же стилях, что отчёт сверки (agent/report.py)."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent.api import KIND_LABELS, KIND_ORDER

from . import models
from .flags import document_flags
from .nomenclature import EXCLUDED
from .passport import NORM_LABELS

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='D9E1F2')
MUTED = Font(name=FONT, size=9, color='808080')
LEVEL_FILL = {
    'g': PatternFill('solid', fgColor='C6EFCE'),
    'y': PatternFill('solid', fgColor='FFEB9C'),
    'r': PatternFill('solid', fgColor='FFC7CE'),
}

# уровень расхождения в терминах agent.api -> цвет и слово
FINDING_LEVELS = {'red': ('r', 'важно'), 'amber': ('y', 'посмотреть'),
                  'ok': ('g', 'сошлось')}
# код расхождения -> что именно проверялось: в таблице это первая колонка,
# по ней замечания фильтруют и сортируют
FINDING_TITLES = {
    'sheet_gap': 'нумерация ведомости',
    'ref_missing': 'прилагаемый документ',
    'ref_cipher': 'шифр документа',
    'spec_sheets': 'листы спецификации',
    'revision_mismatch': 'номера изменений',
    'norm': 'нормативная база',
    'symbol_unused': 'условные обозначения',
    'unreadable': 'читаемость листов',
}
REF_KINDS = {'referenced': 'ссылочный', 'attached': 'прилагаемый',
             'volume': 'комплект раздела'}
DECISIONS = {models.AUTO: '', models.TAKE: 'взято', models.SKIP: 'снято'}
REMARK_STATUS = {models.OPEN: 'в работе', models.SENT: 'передано бюро',
                 models.DISMISSED: 'снято'}


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    return ws


def workbook_bytes(project, submission, documents, rows, totals):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Сводка')
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 60
    lines = [
        ('Объект', project.name),
        ('Шифр', project.code),
        ('Проектное бюро', project.bureau),
        ('Подача', submission.label if submission else ''),
        ('Томов в комплекте', len(documents)),
        ('Листов всего', sum(d.pages_total or 0 for d in documents)),
        ('Позиций в номенклатуре', totals['total']),
        ('— без марки (машинной сверке не поддаются)', totals['no_mark']),
        ('— дубли между разделами', totals['duplicates']),
        ('— исключено изменениями', totals['excluded']),
    ]
    for i, (k, v) in enumerate(lines, 1):
        ws.cell(i, 1, k).font = Font(name=FONT, bold=(i == 1), size=10)
        ws.cell(i, 2, v).font = Font(name=FONT, size=10)
    note = ('Приёмка комплекта: состав томов и сводная номенклатура по данным '
            'спецификаций. Сверка с чертежами не выполнялась. Документ не '
            'является заключением экспертизы.')
    c = ws.cell(len(lines) + 2, 1, note)
    c.font = Font(name=FONT, italic=True, size=9)
    c.alignment = Alignment(wrap_text=True)

    kinds = [k for k in KIND_ORDER
             if any((d.kind_counts or {}).get(k) for d in documents)]
    ws = _sheet(wb, 'Состав комплекта',
                ['Шифр', 'Раздел', 'Изменения', 'Файл', 'Листов']
                + [KIND_LABELS[k] for k in kinds]
                + ['Готовность', 'Флаги'],
                [26, 28, 14, 40, 9] + [13] * len(kinds) + [14, 60])
    for d in documents:
        fl = document_flags(d.capabilities or {}, d.section)
        level = max((f['level'] for f in fl), key=lambda x: 'gyr'.index(x)) if fl else 'g'
        row = ([d.cipher or d.filename, d.section_label or d.section, d.revision,
                d.filename, d.pages_total or 0]
               + [(d.kind_counts or {}).get(k, 0) for k in kinds]
               + [{'g': 'готов', 'y': 'с оговоркой', 'r': 'требует глаз'}[level],
                  '; '.join(f['label'] for f in fl)])
        ws.append(row)
        i = ws.max_row
        for j in range(1, len(row) + 1):
            cell = ws.cell(i, j)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(wrap_text=(j in (4, len(row))), vertical='top')
        ws.cell(i, len(row) - 1).fill = LEVEL_FILL[level]

    ws = _sheet(wb, 'Номенклатура',
                ['Марка', 'Наименование', 'Ед.', 'Кол-во', 'Разделы', 'Томов',
                 'Листы спец.', 'Отметки'],
                [28, 60, 8, 11, 16, 8, 20, 34])
    for r in rows:
        ws.append([r.mark or '—', r.name, r.unit, r.qty if r.qty else None,
                   ', '.join(r.sections), len(r.documents),
                   ', '.join(str(p) for p in sorted(r.pages)[:12]),
                   '; '.join(r.flags)])
        i = ws.max_row
        for j in range(1, 9):
            cell = ws.cell(i, j)
            cell.font = MUTED if EXCLUDED in r.flags else Font(name=FONT, size=9)
            cell.alignment = Alignment(wrap_text=(j == 2), vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------- паспорт тома и план проверки

def _row(ws, values, wrap=(), size=9, fill=None, fill_col=None):
    ws.append(list(values))
    i = ws.max_row
    for j in range(1, len(values) + 1):
        cell = ws.cell(i, j)
        cell.font = Font(name=FONT, size=size)
        cell.alignment = Alignment(wrap_text=(j in wrap), vertical='top')
    if fill and fill_col:
        ws.cell(i, fill_col).fill = fill
    return i


def _text(ws, row, col, value, italic=False, bold=False, size=10):
    c = ws.cell(row, col, value)
    c.font = Font(name=FONT, size=size, bold=bold, italic=italic)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def passport_workbook_bytes(project, submission, doc, psp, plan, users=None,
                            match_rows=(), remarks=()):
    """Паспорт тома и план проверки одной книгой.

    Экран показывает то же самое, но в переписке с бюро нужен файл, который
    можно переслать и в котором можно сортировать. Поэтому книга повторяет
    порядок вкладки: сначала расхождения, потом факты, последним — план.

    Картинки условных обозначений в книгу не кладутся: их вставка тянет
    Pillow в зависимости прода ради колонки, которую всё равно смотрят
    в портале. В листе остаются код, подпись и лист-источник.
    """
    users = users or {}
    wb = Workbook()
    wb.remove(wb.active)
    stats = plan.get('stats') or {}
    rows = plan.get('all_rows') or []

    # --- сводка
    ws = wb.create_sheet('Сводка')
    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 64
    lines = [
        ('Объект', project.name),
        ('Шифр объекта', project.code),
        ('Проектное бюро', project.bureau),
        ('Подача', submission.label if submission else ''),
        ('Том', doc.cipher or doc.filename),
        ('Раздел', doc.section_label or doc.section),
        ('Изменения', doc.revision),
        ('Файл', doc.filename),
        ('Страниц в файле', doc.pages_total or 0),
        ('Листов в ведомости', sum(1 for s in psp['sheets'] if not s['missing'])),
        ('Расхождений найдено', len(psp['findings'])),
        ('— из них важных', sum(1 for f in psp['findings'] if f.get('level') == 'red')),
        ('Проверок сошлось', psp['passed']),
        ('Нормативов в общих указаниях', len(psp['norms'])),
        ('— требуют внимания', len(psp['problem_norms'])),
        ('Условных обозначений', len(psp['symbols'])),
        ('Позиций спецификации', stats.get('total', 0)),
        ('— машина предлагает проверить', stats.get('proposed', 0)),
        ('— отобрано в проверку', stats.get('included', 0)),
    ]
    ms = doc.match_stats or {}
    if match_rows and not ms.get('error'):
        lines += [
            ('Сверка с чертежами, марок', ms.get('rows', len(match_rows))),
            ('— расхождений', ms.get('problems', 0)),
            ('— под вопросом', ms.get('doubts', 0)),
            ('— сошлось', ms.get('matched', 0)),
            ('— позиций машине не проверить', ms.get('uncheckable', 0)),
        ]
    if remarks:
        lines += [
            ('Замечаний', len(remarks)),
            ('— в работе', sum(1 for r in remarks if r.status == 'open')),
            ('— передано бюро', sum(1 for r in remarks if r.status == 'sent')),
            ('— снято экспертом', sum(1 for r in remarks if r.status == 'dismissed')),
        ]
    lines += [('Выгружено', datetime.now().strftime('%d.%m.%Y %H:%M'))]
    for i, (k, v) in enumerate(lines, 1):
        _text(ws, i, 1, k, bold=(i == 1))
        _text(ws, i, 2, v)
    if plan.get('plan') is not None:
        state = 'зафиксирован' if plan.get('frozen') else 'черновик'
        _text(ws, len(lines) + 1, 1, 'План проверки')
        _text(ws, len(lines) + 1, 2, f"№{plan['plan'].version}, {state}")
    _text(ws, len(lines) + 3, 1,
          'Паспорт тома собран разбором PDF: сверены между собой общие данные, '
          'ведомость чертежей и спецификация. Сверка с чертежами не выполнялась. '
          'Класс позиции предложен машиной, отбор в проверку делает эксперт. '
          'Документ не является заключением экспертизы.', italic=True, size=9)

    # --- расхождения
    ws = _sheet(wb, 'Расхождения', ['Уровень', 'Что проверялось', 'Расхождение', 'Листы'],
                [14, 26, 96, 20])
    for f in psp['findings']:
        level, word = FINDING_LEVELS.get(f.get('level'), ('', ''))
        _row(ws, [word, FINDING_TITLES.get(f.get('code'), f.get('code', '')),
                  f.get('text', ''),
                  ', '.join(str(s) for s in (f.get('sheets') or []))],
             wrap=(3,), fill=LEVEL_FILL.get(level), fill_col=1)
    if not psp['findings']:
        _row(ws, ['сошлось', 'все проверки', 'Расхождений в составе не найдено', ''])

    # --- ведомость чертежей: пропуски отдельной строкой, как на экране
    ws = _sheet(wb, 'Ведомость чертежей',
                ['Лист', 'Наименование', 'Изм.', 'Отметка'], [8, 88, 14, 18])
    for s in psp['sheets']:
        if s['missing']:
            _row(ws, [s['no'], 'в ведомости пропущен', '', ''],
                 wrap=(2,), fill=LEVEL_FILL['r'], fill_col=2)
            continue
        _row(ws, [s['no'], s['title'],
                  ', '.join(str(r) for r in (s['revisions'] or [])), s['mark']],
             wrap=(2,))

    # --- документы: ссылочные, прилагаемые, соседние комплекты
    ws = _sheet(wb, 'Документы',
                ['Вид', 'Шифр', 'Наименование', 'Листов объявлено', 'В подаче', 'Стр.'],
                [20, 34, 62, 16, 14, 8])
    for r in list(psp['refs']) + list(psp['volumes']):
        present = 'да' if r.present else ('' if r.kind == 'referenced' else 'нет')
        fill = LEVEL_FILL['r'] if present == 'нет' else None
        _row(ws, [REF_KINDS.get(r.kind, r.kind), r.code, r.title,
                  r.sheets_declared or '', present, r.src_page or ''],
             wrap=(3,), fill=fill, fill_col=5)

    # --- нормативная база
    ws = _sheet(wb, 'Нормативы',
                ['Обозначение', 'Наименование', 'Статус', 'Чем заменён', 'Примечание'],
                [26, 66, 16, 34, 40])
    for n in psp['norms']:
        fill = LEVEL_FILL.get(n.level) if not n.contextual and n.level else None
        _row(ws, [n.code, n.title, NORM_LABELS.get(n.status, n.status),
                  n.replaced_by, n.note + (' (упомянут в названии другого документа)'
                                           if n.contextual else '')],
             wrap=(2, 5), fill=fill, fill_col=3)

    # --- регистрация изменений
    ws = _sheet(wb, 'Изменения',
                ['Изм.', 'Листы', 'Содержание', 'Документ-основание', 'Стр.'],
                [8, 22, 88, 30, 8])
    for r in psp['revisions']:
        _row(ws, [r.number or '', r.sheets, r.content, r.doc_code or r.basis,
                  r.src_page or ''], wrap=(3,))

    # --- условные обозначения
    ws = _sheet(wb, 'Обозначения',
                ['Код', 'Подпись', 'Есть в спецификации', 'Лист'], [22, 76, 20, 8])
    for s in psp['symbols']:
        _row(ws, [s.code, s.name, 'да' if s.used else 'нет', s.page or ''],
             wrap=(2,), fill=None if s.used else LEVEL_FILL['y'], fill_col=3)

    # --- план проверки
    ws = _sheet(wb, 'План проверки',
                ['В проверку', 'Класс', 'Балл', 'Поз.', 'Марка', 'Наименование',
                 'Ед.', 'Кол-во', 'Лист', 'Почему предложено', 'Чем проверяется',
                 'Решение эксперта', 'Кто', 'Комментарий'],
                [12, 8, 8, 8, 30, 60, 8, 12, 8, 46, 24, 16, 18, 40])
    for i in rows:
        reasons = '; '.join(r.get('text', '') for r in (i.reasons or []))
        _row(ws, ['да' if i.included else '', i.cls, i.score, i.pos, i.mark or '—',
                  i.name, i.unit, i.qty if i.qty else None, i.page or '',
                  reasons, ', '.join(i.verifiable_by or []),
                  DECISIONS.get(i.decision, ''),
                  users.get(i.decided_by, ''), i.comment],
             wrap=(6, 10, 14),
             fill=LEVEL_FILL['g'] if i.included else None, fill_col=1)

    # --- замечания идут первым листом после сводки: это то, что уходит бюро
    if remarks:
        ws = _sheet(wb, 'Замечания',
                    ['№', 'Статус', 'Уровень', 'Откуда', 'Предмет',
                     'Формулировка', 'Что нашла машина', 'Листы', 'Эксперт'],
                    [6, 16, 14, 12, 30, 90, 60, 20, 22])
        for n, r in enumerate(remarks, 1):
            level, word = FINDING_LEVELS.get(r.level, ('', ''))
            _row(ws, [n, REMARK_STATUS.get(r.status, r.status), word,
                      'сверка' if r.source == 'match' else 'паспорт',
                      r.subject, r.text, r.evidence,
                      ', '.join(str(p) for p in (r.sheets or [])),
                      users.get(r.author_id, '')],
                 wrap=(6, 7), fill=LEVEL_FILL.get(level), fill_col=3)
        wb.move_sheet('Замечания', offset=-(len(wb.sheetnames) - 2))

    # --- сверка с чертежами: лист появляется, только если она запускалась
    if match_rows:
        ws = _sheet(wb, 'Сверка',
                    ['В плане', 'Марка', 'Наименование по спецификации', 'Ед.',
                     'Спец.', 'Планы', 'Планы без множителя', 'Схемы',
                     'Схемы без множителя', 'Точный источник', 'Статус',
                     'Чем сверено', 'Листы спец.', 'Листы планов', 'Листы схем'],
                    [10, 30, 56, 8, 11, 11, 16, 11, 16, 14, 26, 24, 16, 20, 20])
        for r in match_rows:
            _row(ws, ['да' if r.in_plan else '',
                      ' · '.join(r.marks or []) or r.mark, r.names, r.unit,
                      r.spec_qty, r.plan_qty, r.plan_raw, r.schema_qty,
                      r.schema_raw, r.exact_qty, r.status, r.source,
                      ', '.join(str(p) for p in (r.spec_pages or [])),
                      ', '.join(str(p) for p in (r.plan_pages or [])[:20]),
                      ', '.join(str(p) for p in (r.schema_pages or [])[:20])],
                 wrap=(3,),
                 fill=LEVEL_FILL.get(FINDING_LEVELS.get(r.level, ('', ''))[0]),
                 fill_col=11)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
